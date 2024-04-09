import os
import unittest
from unittest.mock import patch, MagicMock, call

from openpyxl import Workbook

from cat_attack.file_manipulation import find_xlsx_files, attack_file

TEST_PARAMS = {"param": "value"}


class TestFileManipulation(unittest.TestCase):
    def setUp(self):
        self.test_dir = "test_dir"
        os.makedirs(self.test_dir, exist_ok=True)
        self.test_file_path = os.path.join(self.test_dir, "test.xlsx")
        self.wb = Workbook()
        self.wb.save(self.test_file_path)

        self.test_file_path_invalid = os.path.join(self.test_dir, "invalid_file.xlsx.txt")
        self.wb = Workbook()
        self.wb.save(self.test_file_path_invalid)

    @patch("os.path.isdir")
    def test_invalid_directory(self, mock_isdir: MagicMock):
        """
        Tests if the function behaves correctly when given an invalid directory path.

        :param mock_isdir: mocked os.path.isdir function
        :type mock_isdir: MagicMock
        """
        mock_isdir.return_value = False
        self.assertFalse(find_xlsx_files("invalid_directory"))

    @patch("os.path.isdir")
    @patch("glob.glob")
    def test_no_xlsx_files_found(self, mock_glob: MagicMock, mock_isdir: MagicMock):
        """
        Tests if the function behaves correctly when there are no valid .xlsx files found in the given directory.

        :param mock_glob: mocked glob.glob function
        :type mock_glob: MagicMock
        :param mock_isdir: mocked os.path.isdir function
        :type mock_isdir: MagicMock
        """
        mock_isdir.return_value = True
        mock_glob.return_value = []
        self.assertFalse(find_xlsx_files("no xlsx files"))

    @patch("builtins.print")
    def test_invalid_files(self, mock_print: MagicMock):
        """
        Tests if the function behaves correctly when one of the found files is invalid.

        :param mock_print: mocked built-in print function
        :type mock_print: MagicMock
        """
        attack_file(self.test_file_path_invalid, TEST_PARAMS)

        mock_print.assert_called_with(f"Failed to open {os.path.basename(self.test_file_path_invalid)}.")

    @patch("builtins.print")
    def test_already_attacked_file(self, mock_print: MagicMock):
        """
        Tests if the function correctly ignores files with an "Important" sheet if ignore_already_attacked has been
        set to True.

        :param mock_print: mocked built-in print function
        :type mock_print: MagicMock
        """
        self.wb.create_sheet("Important")
        self.wb.save(self.test_file_path)

        attack_file(self.test_file_path, TEST_PARAMS, ignore_already_attacked=True)

        expected_call = [call(f"{os.path.basename(self.test_file_path)} has already been attacked 🙀")]
        mock_print.assert_has_calls(expected_call)

    def tearDown(self):
        os.remove(self.test_file_path)
        os.remove(self.test_file_path_invalid)
        os.rmdir(self.test_dir)


if __name__ == "__main__":
    unittest.main()

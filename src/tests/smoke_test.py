import os
import unittest
from unittest.mock import patch, call, MagicMock

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cat_attack.attack import cat_attack


class SmokeTest(unittest.TestCase):
    def setUp(self):
        self.test_dir = "test_dir"
        os.makedirs(self.test_dir, exist_ok=True)
        self.test_file_path = os.path.join(self.test_dir, "test.xlsx")
        wb = Workbook()
        wb.save(self.test_file_path)

        self.test_file_path_already_attacked = os.path.join(
            self.test_dir, "test_already_attacked.xlsx"
        )
        wb_already_attacked = Workbook()
        wb_already_attacked.create_sheet("Important")
        wb_already_attacked.save(self.test_file_path_already_attacked)

    @patch("builtins.print")
    def test_cat_attack_repeat(self, mock_print: MagicMock):
        """
        Tests if the function correctly identifies whether a file has already been attacked when the
        ignore_already_attacked flag has been set to True.
        """
        cat_attack(self.test_dir, ignore_already_attacked=True)
        calls = [
            call(f"{os.path.basename(self.test_file_path_already_attacked)} has already been attacked 🙀"),
            call("Cats finished attacking all your files 😼"),
        ]
        mock_print.assert_has_calls(calls)

    def test_cat_attack_sheet(self):
        """
        Tests if the function correctly adds an "Important" sheet to the .xlsx file.
        """
        cat_attack(self.test_dir)
        wb: Workbook = load_workbook(self.test_file_path)
        self.assertIn("Important", wb.sheetnames)

    def test_cat_attack_image(self):
        """
        Tests if the function correctly adds an image to the "Important" sheet.
        """
        cat_attack(self.test_dir)
        wb: Workbook = load_workbook(self.test_file_path)
        important_sheet: Worksheet = wb["Important"]
        self.assertTrue(important_sheet._images)

    def tearDown(self):
        os.remove(self.test_file_path)
        os.remove(self.test_file_path_already_attacked)
        os.rmdir(self.test_dir)


if __name__ == "__main__":
    unittest.main()

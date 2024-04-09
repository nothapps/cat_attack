import glob
import os

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException

from cat_attack.cat_image import download_cat_pic

SEARCH_URL = "http://api.thecatapi.com/v1/images/search"


def find_xlsx_files(folder_path: str) -> list[str]:
    """
    Finds all .xlsx files in a given directory and returns a list of found file paths. If the directory path isn't
    valid or contains no .xlsx files it returns an empty list.

    :param folder_path: path to a directory presumably containing .xlsx files
    :type folder_path: str
    :return: a list of found file paths
    :rtype: list[str]
    """
    if not os.path.isdir(folder_path):
        print(f"Folder path: {folder_path} you provided isn't valid.")
        return []
    found_files = glob.glob(folder_path + "/**/*.xlsx", recursive=True)

    if not found_files:
        print("No .xlsx files found in this folder.")
        return []
    else:
        return found_files


def attack_file(file: str, search_params: dict[str, str], ignore_already_attacked: bool = False) -> None:
    """
    Adds an "Important" sheet into each file and inserts a random cat image. If
    ignore_already_attacked is set to True then the function will ignore all the .xlsx files that already have an
    "Important" sheet. If the file cannot be opened or the image can't be downloaded it prints an error message.

    :param file: file path
    :type file: str
    :param search_params: search parameters used to modify the query string in the url
    :type search_params: dict[str, str]
    :param ignore_already_attacked: whether a file has already been attacked (has an "Important" sheet), optional
    :type ignore_already_attacked: bool
    :return: nothing
    :rtype: None
    """
    try:
        current_file: Workbook = load_workbook(file)
    except InvalidFileException:
        print(f"Failed to open {os.path.basename(file)}.")
    else:
        if ignore_already_attacked and "Important" in current_file.sheetnames:
            print(f"{os.path.basename(file)} has already been attacked 🙀")
            return

        important_sheet: Worksheet = current_file.create_sheet("Important")
        cat_image = download_cat_pic(SEARCH_URL, search_params)

        if cat_image is None:
            print(f"Unable to download a cat image for {os.path.basename(file)}")
            return

        important_sheet.add_image(cat_image, "A1")
        current_file.save(file)
